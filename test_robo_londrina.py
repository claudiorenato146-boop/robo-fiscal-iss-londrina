from __future__ import annotations

import sys
import tempfile
import unittest
import zipfile
from datetime import date
from pathlib import Path
from types import SimpleNamespace


PASTA_PROJETO = Path(__file__).resolve().parent
if str(PASTA_PROJETO) not in sys.path:
    sys.path.insert(0, str(PASTA_PROJETO))

from robo_londrina import (  # noqa: E402
    ClienteLondrina,
    PastaDominio,
    aguardar_portal,
    avaliar_zips_existentes,
    carregar_clientes_londrina,
    contar_xmls_no_zip,
    dividir_periodo_em_blocos,
    descobrir_pastas_de_cliente_dominio,
    limitar_periodo_ate_hoje,
    obter_pasta_competencia,
    obter_pasta_competencia_relatorio,
    salvar_relatorio_teste_geral,
    separar_codigos,
    selecionar_contribuinte,
)
from controle_resultados import atualizar_controle, inicializar_controle  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402


class CampoFalso:
    def wait_for(self, **_kwargs):
        return None

    def get_attribute(self, nome: str):
        if nome == "data-value":
            return "12345678000195"
        return None


class ProcessamentoFalso:
    def wait_for(self, **_kwargs):
        return None

    def count(self) -> int:
        return 1

    def nth(self, _indice: int) -> "ProcessamentoFalso":
        return self


class PaginaApexFalsa:
    def __init__(self):
        self.avaliacoes = 0

    def locator(self, seletor: str):
        if seletor == ".u-Processing":
            return ProcessamentoFalso()
        return CampoFalso()

    def evaluate(self, _codigo: str, argumento=None):
        self.avaliacoes += 1
        if argumento is not None:
            return {"ok": True, "valor": argumento["cnpj"]}
        return "12345678000195"

    def wait_for_timeout(self, _milissegundos: int):
        return None


class TestePeriodo(unittest.TestCase):
    def test_competencia_corrente_para_em_hoje(self) -> None:
        inicio, fim = limitar_periodo_ate_hoje(
            date(2026, 7, 1), date(2026, 7, 31), date(2026, 7, 30)
        )
        self.assertEqual(inicio, date(2026, 7, 1))
        self.assertEqual(fim, date(2026, 7, 30))

    def test_periodo_futuro_e_bloqueado(self) -> None:
        with self.assertRaises(SystemExit):
            limitar_periodo_ate_hoje(
                date(2026, 8, 1), date(2026, 8, 31), date(2026, 7, 30)
            )

    def test_divisao_respeita_sessenta_dias(self) -> None:
        blocos = dividir_periodo_em_blocos(
            date(2026, 1, 1), date(2026, 3, 15)
        )
        self.assertEqual(blocos[0], (date(2026, 1, 1), date(2026, 3, 1)))
        self.assertEqual(blocos[1], (date(2026, 3, 2), date(2026, 3, 15)))


class TesteProcessamento(unittest.TestCase):
    def test_espera_portal_usa_limite_longo(self) -> None:
        pagina = PaginaApexFalsa()
        processamento = ProcessamentoFalso()
        chamadas = []
        processamento.wait_for = lambda **kwargs: chamadas.append(kwargs)
        pagina.locator = lambda _seletor: processamento

        aguardar_portal(pagina, "teste")

        self.assertEqual(chamadas[0]["state"], "hidden")
        self.assertEqual(chamadas[0]["timeout"], 120_000)

    def test_csv_com_bom_e_aceito(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            caminho = Path(temporaria) / "clientes.csv"
            caminho.write_text(
                "codigo_cliente,cnpj,apelido,pasta_dominio,razao_social\n"
                "19,99.900.004/0001-93,CLIENTE,19-CLIENTE,CLIENTE TESTE\n",
                encoding="utf-8-sig",
            )

            clientes = carregar_clientes_londrina(str(caminho))

            self.assertEqual(len(clientes), 1)
            self.assertEqual(clientes[0].codigo, "19")
            self.assertEqual(clientes[0].cnpj, "99900004000193")

    def test_pasta_com_espaco_antes_do_hifen_e_reconhecida(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            pasta = Path(temporaria) / "544 - CLINICA DEMO"
            pasta.mkdir()
            mapa = descobrir_pastas_de_cliente_dominio(Path(temporaria))
            self.assertEqual(mapa["544"].caminho, pasta)

    def test_pasta_duplicada_usa_o_nome_cadastrado_no_csv(self) -> None:
        """Evita pendência quando há pasta antiga e atual para o mesmo código."""
        with tempfile.TemporaryDirectory() as temporaria:
            raiz = Path(temporaria)
            antiga = raiz / "544-CLINICA DEMONSTRACAO"
            correta = raiz / "544-CLINICA DEMONSTRACAO LTDA"
            antiga.mkdir()
            correta.mkdir()
            cliente = ClienteLondrina(
                "544", "12345678000195", "CLINICA DEMO", correta.name
            )

            mapa = descobrir_pastas_de_cliente_dominio(raiz, {"544": cliente})

            self.assertEqual(mapa["544"].caminho, correta)

    def test_planilha_e_atualizada_empresa_por_empresa(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            cliente = ClienteLondrina(
                "28", "99900003000149", "OFICINA FICTICIA DE", "", "OFICINA FICTICIA DE AUTOPECAS LTDA"
            )
            caminho = inicializar_controle([cliente], "072026", Path(temporaria))
            resultado = SimpleNamespace(
                codigo_cliente="28",
                nome="OFICINA FICTICIA DE AUTOPECAS LTDA",
                cnpj="99900003000149",
                emitidos="X",
                recebidos="OK",
                relatorio="OK",
                detalhe="Emitidos sem movimento; recebidos com 3 XMLs",
            )
            atualizar_controle(caminho, "072026", resultado)
            ws = load_workbook(caminho)["072026"]
            linha = next(i for i in range(2, ws.max_row + 1) if str(ws.cell(i, 1).value) == "28")
            self.assertEqual(ws.cell(linha, 4).value, "X")
            self.assertEqual(ws.cell(linha, 5).value, "OK")
            self.assertIn("3 XMLs", ws.cell(linha, 6).value)

    def test_nova_competencia_cria_nova_aba_sem_apagar_a_anterior(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            cliente = ClienteLondrina(
                "28", "99900003000149", "OFICINA FICTICIA DE", "", "OFICINA FICTICIA DE AUTOPECAS LTDA"
            )
            raiz = Path(temporaria)
            caminho = inicializar_controle([cliente], "072026", raiz)
            inicializar_controle([cliente], "082026", raiz)
            workbook = load_workbook(caminho)
            self.assertIn("072026", workbook.sheetnames)
            self.assertIn("082026", workbook.sheetnames)

    def test_aba_vazia_com_competencia_digitada_errada_e_reaproveitada(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            raiz = Path(temporaria)
            workbook = Workbook()
            ws = workbook.active
            ws.title = "072027"
            ws.append(["N°", "RAZÃO SOCIAL", "CNPJ", "XML - PRESTADOS", "XML - TOMADOS", "OBSERVAÇÃO"])
            workbook.save(raiz / "PLANILHA RELATORIO MENSAL.xlsx")
            cliente = ClienteLondrina(
                "28", "99900003000149", "OFICINA FICTICIA DE", "", "OFICINA FICTICIA DE AUTOPECAS LTDA"
            )
            caminho = inicializar_controle([cliente], "072026", raiz)
            final = load_workbook(caminho)
            self.assertIn("072026", final.sheetnames)
            self.assertNotIn("072027", final.sheetnames)

    def test_zip_com_xml_na_raiz_tem_movimento(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            caminho = Path(temporaria) / "notas.zip"
            with zipfile.ZipFile(caminho, "w") as zf:
                zf.writestr("nota.xml", "<nfe/>")
            self.assertEqual(contar_xmls_no_zip(caminho), 1)
            self.assertEqual(avaliar_zips_existentes([caminho])[:2], ("OK", 1))

    def test_zip_com_xml_em_subpasta_tambem_tem_movimento(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            caminho = Path(temporaria) / "notas.zip"
            with zipfile.ZipFile(caminho, "w") as zf:
                zf.writestr("pasta/NOTA.XML", "<nfe/>")
                zf.writestr("pasta/leia-me.txt", "ignorar")
            self.assertEqual(contar_xmls_no_zip(caminho), 1)

    def test_zip_valido_sem_xml_e_sem_movimento_nao_e_erro(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            caminho = Path(temporaria) / "vazio.zip"
            with zipfile.ZipFile(caminho, "w") as zf:
                zf.writestr("resumo.txt", "sem notas")
            self.assertEqual(contar_xmls_no_zip(caminho), 0)
            self.assertEqual(avaliar_zips_existentes([caminho])[:2], ("X", 0))

    def test_arquivo_corrompido_e_falha_tecnica(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            caminho = Path(temporaria) / "quebrado.zip"
            caminho.write_text("não é zip", encoding="utf-8")
            with self.assertRaises(RuntimeError):
                contar_xmls_no_zip(caminho)

    def test_zip_valido_substitui_uma_copia_antiga_corrompida(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            quebrado = Path(temporaria) / "notas.zip"
            valido = Path(temporaria) / "notas__v2.zip"
            quebrado.write_text("erro antigo", encoding="utf-8")
            with zipfile.ZipFile(valido, "w") as zf:
                zf.writestr("nota.xml", "<nfe/>")
            self.assertEqual(avaliar_zips_existentes([quebrado, valido])[:2], ("OK", 1))

    def test_competencia_nao_e_criada_antes_do_download(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            pasta_cliente = Path(temporaria) / "10-CLIENTE"
            pasta_cliente.mkdir()
            cliente = ClienteLondrina("10", "12345678000195", "CLIENTE", "")
            pastas = {
                "10": PastaDominio("10", pasta_cliente.name, pasta_cliente)
            }
            competencia = obter_pasta_competencia(cliente, pastas, "072026")
            self.assertEqual(competencia, pasta_cliente / "072026")
            self.assertFalse(competencia.exists())

    def test_emitidas_e_recebidas_resolvem_para_raizes_diferentes(self) -> None:
        """
        Regressão da mudança de 03/08/2026: emitidas e recebidas passaram
        a ir para pastas raiz DIFERENTES (Importação vs Importação -
        Tomados). Confirma que obter_pasta_competencia, usado com cada
        mapa de pastas separadamente, aponta para lugares diferentes —
        mesmo cliente, mesma competência, destino diferente.
        """
        with tempfile.TemporaryDirectory() as temporaria:
            raiz = Path(temporaria)
            pasta_cliente_emitidas = raiz / "importacao" / "10-CLIENTE"
            pasta_cliente_tomados = raiz / "importacao_tomados" / "10-CLIENTE"
            pasta_cliente_emitidas.mkdir(parents=True)
            pasta_cliente_tomados.mkdir(parents=True)

            cliente = ClienteLondrina("10", "12345678000195", "CLIENTE", "")
            mapa_emitidas = {"10": PastaDominio("10", "10-CLIENTE", pasta_cliente_emitidas)}
            mapa_tomados = {"10": PastaDominio("10", "10-CLIENTE", pasta_cliente_tomados)}

            destino_emitidas = obter_pasta_competencia(cliente, mapa_emitidas, "072026")
            destino_tomados = obter_pasta_competencia(cliente, mapa_tomados, "072026")

            self.assertEqual(destino_emitidas, pasta_cliente_emitidas / "072026")
            self.assertEqual(destino_tomados, pasta_cliente_tomados / "072026")
            self.assertNotEqual(destino_emitidas, destino_tomados)

    def test_relatorio_reusa_pasta_existente_do_cliente(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            raiz = Path(temporaria)
            pasta_existente = raiz / "1-ON4"
            pasta_existente.mkdir()
            cliente = ClienteLondrina("1", "99900001000150", "PADARIA DO EXEMPLO", "")
            mapa = {"1": PastaDominio("1", "1-ON4", pasta_existente)}

            resultado = obter_pasta_competencia_relatorio(cliente, mapa, raiz, "072026")

            self.assertEqual(resultado, pasta_existente / "072026")
            self.assertFalse(resultado.exists())  # a subpasta de competência não é criada aqui

    def test_relatorio_cria_pasta_do_cliente_se_nao_existir(self) -> None:
        """
        Regra EXPLICITAMENTE diferente das pastas de importação: aqui o
        robô PODE criar a pasta do cliente (pedido do usuário, 03/08/2026).
        """
        with tempfile.TemporaryDirectory() as temporaria:
            raiz = Path(temporaria)
            cliente = ClienteLondrina("99", "12345678000199", "CLIENTE NOVO", "")

            resultado = obter_pasta_competencia_relatorio(cliente, {}, raiz, "072026")

            pasta_criada = raiz / "99-CLIENTE NOVO"
            self.assertTrue(pasta_criada.exists())
            self.assertEqual(resultado, pasta_criada / "072026")

    def test_relatorio_sem_apelido_usa_so_codigo(self) -> None:
        with tempfile.TemporaryDirectory() as temporaria:
            raiz = Path(temporaria)
            cliente = ClienteLondrina("50", "98765432000188", "", "")

            obter_pasta_competencia_relatorio(cliente, {}, raiz, "072026")

            self.assertTrue((raiz / "50").exists())

    def test_selecao_direta_pela_api_apex(self) -> None:
        cliente = ClienteLondrina("10", "12345678000195", "CLIENTE", "")
        pagina = PaginaApexFalsa()
        self.assertTrue(selecionar_contribuinte(pagina, cliente))
        self.assertEqual(pagina.avaliacoes, 2)

    def test_selecao_funciona_com_multiplos_indicadores_de_carregamento(self) -> None:
        """
        Regressão do erro real de 03/08/2026: 'strict mode violation:
        locator(".u-Processing") resolved to 2 elements' — a tela de
        Apurações usa a mesma classe CSS para seu próprio indicador,
        então pode haver mais de um ao mesmo tempo. Simula 2 elementos
        para confirmar que não quebra mais.
        """
        cliente = ClienteLondrina("10", "12345678000195", "CLIENTE", "")
        pagina = PaginaApexFalsa()
        pagina.locator(".u-Processing").count = lambda: 2  # simula 2 elementos
        self.assertTrue(selecionar_contribuinte(pagina, cliente))


class TesteFuncionalidadesNovas(unittest.TestCase):
    """Cobre os recursos adicionados em 05/08/2026."""

    def test_separar_codigos_por_ponto(self) -> None:
        self.assertEqual(
            separar_codigos("545.582.548.547"),
            ["545", "582", "548", "547"],
        )

    def test_separar_codigos_aceita_virgula_e_espaco(self) -> None:
        self.assertEqual(separar_codigos("545, 582; 548 547"), ["545", "582", "548", "547"])

    def test_separar_codigos_remove_repetidos_mantendo_ordem(self) -> None:
        self.assertEqual(separar_codigos("10.10.20.10.30"), ["10", "20", "30"])

    def test_separar_codigos_vazio(self) -> None:
        self.assertEqual(separar_codigos(""), [])
        self.assertEqual(separar_codigos("   "), [])

    def test_relatorio_teste_geral_gera_txt_com_resumo(self) -> None:
        linhas = [
            {"codigo": "1", "nome": "PADARIA DO EXEMPLO", "cnpj": "99900001000150",
             "perfil": "OK", "detalhe": "seleção e abas OK"},
            {"codigo": "50", "nome": "EMPRESA SEM PROC", "cnpj": "12345678000195",
             "perfil": "SEM ACESSO", "detalhe": "não apareceu na lista"},
        ]
        import os
        cwd = os.getcwd()
        with tempfile.TemporaryDirectory() as temporaria:
            os.chdir(temporaria)
            try:
                caminho = salvar_relatorio_teste_geral(linhas)
                conteudo = caminho.read_text(encoding="utf-8")
            finally:
                os.chdir(cwd)
        self.assertIn("RELATÓRIO DE DIAGNÓSTICO", conteudo)
        self.assertIn("Total de empresas testadas: 2", conteudo)
        self.assertIn("SEM ACESSO", conteudo)
        self.assertIn("EMPRESA SEM PROC", conteudo)
        # O grupo SEM ACESSO deve vir antes do grupo OK (prioriza atenção).
        self.assertLess(conteudo.index("SEM ACESSO  (1)"), conteudo.index("OK  (1)"))


if __name__ == "__main__":
    unittest.main()
