"""Atualização incremental da planilha mensal única do setor fiscal."""

from __future__ import annotations

import os
import re
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


NOME_ARQUIVO = "PLANILHA RELATORIO MENSAL.xlsx"
CABECALHOS = [
    "N°", "RAZÃO SOCIAL", "CNPJ", "XML - PRESTADOS", "XML - TOMADOS", "OBSERVAÇÃO"
]


def formatar_cnpj(cnpj: str) -> str:
    digitos = "".join(c for c in cnpj if c.isdigit()).zfill(14)
    return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"


def _salvar_atomico(workbook, caminho: Path) -> None:
    temporario = caminho.with_name(f".{caminho.stem}.temporario.xlsx")
    workbook.save(temporario)
    os.replace(temporario, caminho)


def _somente_cabecalho(ws) -> bool:
    return ws.max_row == 1 and [ws.cell(1, c).value for c in range(1, 7)] == CABECALHOS


def _nome_tabela(competencia: str) -> str:
    return "Tabela" + re.sub(r"\W", "", competencia)


def _formatar_aba(ws, ultima_linha: int) -> None:
    azul = "1F4E78"
    azul_claro = "D9EAF7"
    borda = Side(style="thin", color="D9E2F3")
    borda_escura = Side(style="thin", color="17365D")

    for coluna, cabecalho in enumerate(CABECALHOS, 1):
        celula = ws.cell(1, coluna, cabecalho)
        celula.fill = PatternFill("solid", fgColor=azul)
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = Border(
            left=borda_escura, right=borda_escura, top=borda_escura, bottom=borda_escura
        )
    ws.row_dimensions[1].height = 28

    for linha in range(2, ultima_linha + 1):
        if linha % 2 == 0:
            for coluna in range(1, 7):
                ws.cell(linha, coluna).fill = PatternFill("solid", fgColor=azul_claro)
        for coluna in range(1, 7):
            ws.cell(linha, coluna).border = Border(
                left=borda, right=borda, top=borda, bottom=borda
            )
            ws.cell(linha, coluna).alignment = Alignment(vertical="center")
        ws.cell(linha, 1).alignment = Alignment(horizontal="center", vertical="center")
        for coluna in (3, 4, 5):
            ws.cell(linha, coluna).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(linha, 2).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(linha, 6).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[linha].height = 22

    larguras = {"A": 9, "B": 43, "C": 21, "D": 18, "E": 18, "F": 58}
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # Reaplica as cores de status em toda a área de dados.
    ws.conditional_formatting._cf_rules.clear()
    faixa = f"D2:E{ultima_linha}"
    ws.conditional_formatting.add(
        faixa,
        CellIsRule(
            operator="equal", formula=['"OK"'],
            fill=PatternFill("solid", fgColor="C6EFCE"),
            font=Font(color="006100", bold=True),
        ),
    )
    ws.conditional_formatting.add(
        faixa,
        CellIsRule(
            operator="equal", formula=['"X"'],
            fill=PatternFill("solid", fgColor="FFF2CC"),
            font=Font(color="9C6500", bold=True),
        ),
    )
    ws.conditional_formatting.add(
        faixa,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("ERRO",D2))'],
            fill=PatternFill("solid", fgColor="FFC7CE"),
            font=Font(color="9C0006", bold=True),
        ),
    )

    referencia = f"A1:F{ultima_linha}"
    if ws.tables:
        tabela = next(iter(ws.tables.values()))
        tabela.ref = referencia
    else:
        tabela = Table(displayName=_nome_tabela(ws.title), ref=referencia)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tabela)


def _obter_ou_criar_aba(workbook, competencia: str):
    if competencia in workbook.sheetnames:
        return workbook[competencia]

    # Corrige automaticamente uma única aba vazia usada como modelo —
    # inclusive o caso recebido em que estava nomeada 072027 em vez de 072026.
    candidatas_vazias = [ws for ws in workbook.worksheets if _somente_cabecalho(ws)]
    if len(candidatas_vazias) == 1:
        candidatas_vazias[0].title = competencia
        return candidatas_vazias[0]

    modelo = workbook[workbook.sheetnames[-1]]
    nova = workbook.copy_worksheet(modelo)
    nova.title = competencia
    if nova.max_row > 1:
        nova.delete_rows(2, nova.max_row - 1)
    # Tabelas não são copiadas pelo openpyxl; se alguma versão as copiar,
    # remove para evitar nomes duplicados e recria na formatação.
    nova._tables.clear()
    return nova


def inicializar_controle(clientes, competencia: str, raiz_relatorios: Path) -> Path:
    """Abre a planilha única e prepara/reutiliza a aba da competência."""
    raiz_relatorios.mkdir(parents=True, exist_ok=True)
    caminho = raiz_relatorios / NOME_ARQUIVO
    modelo = Path(__file__).with_name(NOME_ARQUIVO)
    if not caminho.exists():
        if not modelo.exists():
            raise FileNotFoundError(f"Modelo da planilha não encontrado: {modelo}")
        shutil.copy2(modelo, caminho)

    workbook = load_workbook(caminho)
    ws = _obter_ou_criar_aba(workbook, competencia)

    anteriores = {}
    for linha in range(2, ws.max_row + 1):
        codigo = ws.cell(linha, 1).value
        if codigo is not None:
            anteriores[str(codigo).strip()] = (
                ws.cell(linha, 4).value or "",
                ws.cell(linha, 5).value or "",
                ws.cell(linha, 6).value or "",
            )
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for linha, cliente in enumerate(clientes, 2):
        prestados, tomados, observacao = anteriores.get(
            str(cliente.codigo), ("", "", "Aguardando processamento")
        )
        ws.cell(linha, 1, str(cliente.codigo))
        ws.cell(linha, 2, cliente.nome_exibicao)
        ws.cell(linha, 3, formatar_cnpj(cliente.cnpj))
        ws.cell(linha, 4, prestados)
        ws.cell(linha, 5, tomados)
        ws.cell(linha, 6, observacao)

    _formatar_aba(ws, len(clientes) + 1)
    _salvar_atomico(workbook, caminho)
    return caminho


def atualizar_controle(caminho: Path, competencia: str, resultado) -> None:
    """Atualiza uma empresa na aba mensal e salva imediatamente."""
    workbook = load_workbook(caminho)
    if competencia not in workbook.sheetnames:
        raise KeyError(f"Aba da competência {competencia} não encontrada")
    ws = workbook[competencia]
    linha = next(
        (
            numero
            for numero in range(2, ws.max_row + 1)
            if str(ws.cell(numero, 1).value).strip() == str(resultado.codigo_cliente)
        ),
        None,
    )
    if linha is None:
        linha = ws.max_row + 1
        ws.cell(linha, 1, str(resultado.codigo_cliente))
        ws.cell(linha, 2, resultado.nome)
        ws.cell(linha, 3, formatar_cnpj(resultado.cnpj))

    ws.cell(linha, 4, resultado.emitidos)
    ws.cell(linha, 5, resultado.recebidos)
    ws.cell(linha, 6, resultado.detalhe or "")
    _formatar_aba(ws, ws.max_row)
    _salvar_atomico(workbook, caminho)

